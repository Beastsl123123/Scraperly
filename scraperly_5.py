import tkinter as tk
from tkinter import ttk, messagebox
import threading
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import webbrowser


class NewsScraper:
    def __init__(self):
        self.sites = [
            {"name": "Al Jazeera", "url": "https://www.aljazeera.com/news/", "tag": "h3", "class_name": "gc__title", "link_tag": "a", "link_class": "gc__title", "base_url": "https://www.aljazeera.com", "site_url": "https://www.aljazeera.com"},
            {"name": "BBC", "url": "https://www.bbc.com/news", "tag": "h3", "class_name": "gs-c-promo-heading__title", "link_tag": "a", "link_class": "gs-c-promo-heading", "base_url": "https://www.bbc.com", "site_url": "https://www.bbc.com"},
            {"name": "Reuters", "url": "https://www.reuters.com/world/", "tag": "h3", "class_name": "story-title", "link_tag": "a", "link_class": "story-title", "base_url": "https://www.reuters.com", "site_url": "https://www.reuters.com"},
            {"name": "CNN", "url": "https://edition.cnn.com/world", "tag": "span", "class_name": "cd__headline-text", "link_tag": "a", "link_class": "container__link", "base_url": "https://edition.cnn.com", "site_url": "https://edition.cnn.com"},
            {"name": "NY Times", "url": "https://www.nytimes.com/section/world", "tag": "h2", "class_name": "css-1j9dxys e1xfvim30", "link_tag": "a", "link_class": "css-1wjnrbv", "base_url": "https://www.nytimes.com", "site_url": "https://www.nytimes.com"},
            {"name": "The Guardian", "url": "https://www.theguardian.com/world", "tag": "div", "class_name": "fc-item__content", "link_tag": "a", "link_class": "u-faux-block-link__overlay js-headline-text", "base_url": "https://www.theguardian.com", "site_url": "https://www.theguardian.com"},
            {"name": "Fox News", "url": "https://www.foxnews.com/world", "tag": "h2", "class_name": "title", "link_tag": "a", "link_class": "title", "base_url": "https://www.foxnews.com", "site_url": "https://www.foxnews.com"},
        ]

        self.importance_levels = {
            "High": {"emoji": "🔴", "color_code": "#DC2626"},  # red-600
            "Medium": {"emoji": "🟡", "color_code": "#CA8A04"},  # yellow-600
            "Low": {"emoji": "🟢", "color_code": "#16A34A"},  # green-600
        }

    def fetch_html(self, url):
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/114.0.0.0 Safari/537.36"
            )
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.content

    def parse_headlines_with_links(self, html, tag, class_name, link_tag, link_class, base_url):
        soup = BeautifulSoup(html, "html.parser")
        results = []
        class_split = class_name.split() if class_name else []
        if class_split:
            headline_elements = soup.find_all(
                tag, class_=lambda c: c and all(cls in c.split() for cls in class_split)
            )
        else:
            headline_elements = soup.find_all(tag)
        for hl in headline_elements:
            text = hl.get_text(strip=True)
            if not text:
                continue
            link = None
            candidate_links = hl.find_all(
                link_tag, class_=lambda c: c and link_class in c.split()
            )
            if candidate_links:
                link = candidate_links[0].get("href")
            else:
                parent = hl.parent
                if parent:
                    p_links = parent.find_all(
                        link_tag, class_=lambda c: c and link_class in c.split()
                    )
                    if p_links:
                        link = p_links[0].get("href")
            if link and not link.startswith("http"):
                if base_url:
                    link = base_url.rstrip("/") + "/" + link.lstrip("/")
            results.append({"text": text, "link": link or ""})
        seen = set()
        unique_results = []
        for item in results:
            if item["text"] not in seen:
                seen.add(item["text"])
                unique_results.append(item)
        return unique_results

    def assign_importance(self, headline):
        high_keywords = [
            "crisis", "war", "breaking", "urgent", "alert",
            "disaster", "emergency", "dead", "attack",
        ]
        medium_keywords = [
            "election", "government", "protest", "policy", "update", "accident",
        ]
        lower_text = headline.lower()
        if any(k in lower_text for k in high_keywords):
            return "High"
        elif any(k in lower_text for k in medium_keywords):
            return "Medium"
        else:
            return "Low"

    def scrape_all_sites(self):
        all_data = []
        for site in self.sites:
            try:
                html = self.fetch_html(site["url"])
                headlines_links = self.parse_headlines_with_links(
                    html,
                    site["tag"],
                    site["class_name"],
                    site["link_tag"],
                    site["link_class"],
                    site["base_url"],
                )
                for hl in headlines_links:
                    importance = self.assign_importance(hl["text"])
                    all_data.append({
                        "Source": site["name"],
                        "Headline": hl["text"],
                        "Link": hl["link"],
                        "Importance": importance,
                        "Emoji": self.importance_levels[importance]["emoji"],
                        "SiteUrl": site["site_url"],
                    })
            except Exception as e:
                print(f"Error fetching/parsing {site['name']}: {e}")
        seen = set()
        unique = []
        for item in all_data:
            if item["Headline"] not in seen:
                seen.add(item["Headline"])
                unique.append(item)
        return unique

    def save_to_excel(self, data, filename):
        df = pd.DataFrame(data)
        df = df[["Source", "Headline", "Link", "Importance", "Emoji"]]
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="News Headlines")
            ws = writer.sheets["News Headlines"]

            header_font = Font(bold=True, size=18, color="444444")
            header_fill = PatternFill("solid", fgColor="F9F9F9")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border_side = Side(style="thin", color="DDDDDD")
            border = Border(
                left=border_side, right=border_side, top=border_side, bottom=border_side
            )

            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            widths = {"A": 20, "B": 80, "C": 50, "D": 15, "E": 8}
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width

            importance_fills = {
                "High": PatternFill("solid", fgColor="FFF4F4"),
                "Medium": PatternFill("solid", fgColor="FFFAE5"),
                "Low": PatternFill("solid", fgColor="EDF8F2"),
            }
            importance_fonts = {
                "High": Font(bold=True, color="9C1C1C"),
                "Medium": Font(bold=True, color="7F6B00"),
                "Low": Font(bold=True, color="166530"),
            }

            for row in range(2, ws.max_row + 1):
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border

                headline_cell = ws.cell(row=row, column=2)
                headline_cell.font = Font(bold=True, italic=True, size=14, color="444444")
                headline_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

                link_cell = ws.cell(row=row, column=3)
                link_val = link_cell.value
                if link_val:
                    link_cell.font = Font(underline="single", color="0563C1")
                    link_cell.hyperlink = link_val
                link_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

                imp_cell = ws.cell(row=row, column=4)
                imp_val = imp_cell.value
                if imp_val in importance_fills:
                    imp_cell.fill = importance_fills[imp_val]
                    imp_cell.font = importance_fonts[imp_val]
                imp_cell.alignment = Alignment(horizontal="center", vertical="center")

                emoji_cell = ws.cell(row=row, column=5)
                emoji_cell.alignment = Alignment(horizontal="center", vertical="center")

                source_cell = ws.cell(row=row, column=1)
                source_cell.font = Font(color="444444")
                source_cell.alignment = Alignment(horizontal="center", vertical="center")


class HeadlinesGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Scraperly News - Vertical Layout")
        self.geometry("960x1080")
        self.minsize(700, 800)
        self.configure(bg="#ffffff")

        self.scraper = NewsScraper()
        self.headlines_grouped = {"High": [], "Medium": [], "Low": []}

        self.container = tk.Frame(self, bg="#ffffff")
        self.container.pack(expand=True, fill="both", padx=20, pady=20)

        # Header
        self.header_label = tk.Label(
            self.container,
            text="Scraperly News 🔥",
            font=("Poppins", 48, "bold"),
            fg="#111827",
            bg="#ffffff",
            pady=10,
        )
        self.header_label.pack()

        self.sub_text = tk.Label(
            self.container,
            text="Grouped by Importance Level | Select a headline and open corresponding news site",
            font=("Poppins", 16),
            fg="#6b7280",
            bg="#ffffff",
            pady=10,
        )
        self.sub_text.pack()

        # Buttons
        buttons_frame = tk.Frame(self.container, bg="#ffffff")
        buttons_frame.pack(pady=15)

        self.scrape_button = ttk.Button(
            buttons_frame,
            text="Scrape Headlines",
            command=self.run_scrape,
            width=20,
        )
        self.scrape_button.pack(side="left", padx=10)

        self.export_button = ttk.Button(
            buttons_frame,
            text="Convert to Excel",
            command=self.export_to_excel,
            width=20,
            state="disabled",
        )
        self.export_button.pack(side="left", padx=10)

        # Importance groups - stacked vertically
        self.group_frames = {}
        self.group_trees = {}
        self.group_buttons = {}

        for level in ["High", "Medium", "Low"]:
            frame = tk.LabelFrame(
                self.container,
                text=f"{level} Importance",
                font=("Poppins", 22, "bold"),
                fg="#111827",
                bg="#f9fafb",
                labelanchor="nw",
                relief="groove",
                bd=2,
                padx=15,
                pady=15,
            )
            frame.pack(fill="both", expand=True, pady=10)

            cols = ("Source", "Headline")
            tree = ttk.Treeview(
                frame,
                columns=cols,
                show="headings",
                selectmode="browse",
                height=10,
            )
            for col in cols:
                tree.heading(col, text=col, anchor="w")
                if col == "Headline":
                    tree.column(col, width=600, anchor="w")
                else:
                    tree.column(col, width=150, anchor="center")
            tree.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            scrollbar.pack(side="right", fill="y")
            tree.configure(yscrollcommand=scrollbar.set)

            btn = ttk.Button(
                frame,
                text="Click here to go to the website",
                command=lambda lvl=level: self.open_selected_link(lvl),
                state="disabled",
                width=30,
            )
            btn.pack(pady=10)

            tree.bind("<<TreeviewSelect>>", lambda e, lvl=level: self.on_tree_selection(e, lvl))

            self.group_frames[level] = frame
            self.group_trees[level] = tree
            self.group_buttons[level] = btn

        self.headlines_data = []

        self.after(100, self.run_scrape)
        self.bind('<Configure>', self.on_resize)

    def on_resize(self, event):
        for level in ["High", "Medium", "Low"]:
            tree = self.group_trees[level]
            frame = self.group_frames[level]
            width = frame.winfo_width() or 700
            tree.column("Source", width=150)
            tree.column("Headline", width=width - 200)

    def run_scrape(self):
        self.title("Scraperly News - Scraping...")
        self.disable_all_buttons()
        threading.Thread(target=self.scrape_thread, daemon=True).start()

    def scrape_thread(self):
        try:
            self.headlines_data = self.scraper.scrape_all_sites()
            if not self.headlines_data:
                self.title("Scraperly News - No headlines found")
                self.clear_all_trees()
                self.export_button.config(state="disabled")
                self.enable_all_buttons()
                return

            grouped = {"High": [], "Medium": [], "Low": []}
            for item in self.headlines_data:
                grouped[item["Importance"]].append(item)
            self.headlines_grouped = grouped

            self.clear_all_trees()

            for level in ["High", "Medium", "Low"]:
                tree = self.group_trees[level]
                for item in grouped[level]:
                    hd = item["Headline"]
                    display_hd = hd if len(hd) <= 110 else hd[:107] + "..."
                    tree.insert("", "end", values=(item["Source"], display_hd))
                self.group_buttons[level].config(state="disabled")

            total_count = sum(len(v) for v in grouped.values())
            self.title(f"Scraperly News - {total_count} headlines loaded")
            self.export_button.config(state="normal")
            self.enable_all_buttons()

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred:\n{e}")
            self.title("Scraperly News - Error during scrape")
            self.export_button.config(state="disabled")
            self.enable_all_buttons()

    def disable_all_buttons(self):
        self.scrape_button.config(state="disabled")
        self.export_button.config(state="disabled")
        for btn in self.group_buttons.values():
            btn.config(state="disabled")

    def enable_all_buttons(self):
        self.scrape_button.config(state="normal")

    def clear_all_trees(self):
        for tree in self.group_trees.values():
            tree.delete(*tree.get_children())

    def on_tree_selection(self, event, level):
        tree = self.group_trees[level]
        btn = self.group_buttons[level]
        if tree.selection():
            btn.config(state="normal")
        else:
            btn.config(state="disabled")

    def open_selected_link(self, level):
        tree = self.group_trees[level]
        selection = tree.selection()
        if not selection:
            messagebox.showinfo("No Selection", "Please select a headline first.")
            return
        idx = tree.index(selection[0])
        if idx >= len(self.headlines_grouped[level]):
            messagebox.showinfo("Error", "Selected item out of range.")
            return
        url = self.headlines_grouped[level][idx].get("SiteUrl", "")
        if url:
            webbrowser.open_new_tab(url)
        else:
            messagebox.showinfo("No Link", "No website link available for this news outlet.")

    def export_to_excel(self):
        if not self.headlines_data:
            messagebox.showinfo("No Data", "No headlines to export.")
            return
        try:
            self.scraper.save_to_excel(self.headlines_data, "news_headlines_styled.xlsx")
            messagebox.showinfo("Export Successful", "Headlines exported successfully to news_headlines_styled.xlsx")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not export file:\n{e}")


def main():
    app = HeadlinesGUI()
    app.mainloop()


if __name__ == "__main__":
    main()








